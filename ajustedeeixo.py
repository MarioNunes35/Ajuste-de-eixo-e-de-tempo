
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from datetime import datetime
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# -----------------------------
# Utils
# -----------------------------

def _parse_time_to_seconds_raw(val, mmss_mode: bool = True):
    """
    Parse a single cell into absolute seconds (not anchored).
    If mmss_mode is True:
        - "mm:ss" stays mm:ss
        - "hh:mm:ss" is interpreted as mm:ss by mapping HH->minutes and MM->seconds (ignores SS)
        - datetime-like or Excel fraction: convert to h,m,s then minutes=hour, seconds=minute
    If mmss_mode is False:
        - Use standard hh:mm:ss / mm:ss semantics.
    """
    if pd.isna(val):
        return None

    # Strings --------------------------------------------------
    if isinstance(val, str):
        s = val.strip()
        if " " in s and ":" in s:
            s = s.split()[-1]
        parts = s.split(":")
        try:
            if len(parts) == 2:
                m = int(float(parts[0]))
                s = int(float(parts[1]))
                return m * 60 + s
            elif len(parts) >= 3:
                h = int(float(parts[0]))
                m = int(float(parts[1]))
                s = int(float(parts[2]))
                if mmss_mode:
                    # reinterpret "h:m:s" as "m:s" taking h as minutes and m as seconds
                    return h * 60 + m
                else:
                    return h * 3600 + m * 60 + s
        except Exception:
            return None
        return None

    # Numbers (Excel fractional day or already seconds) --------
    if isinstance(val, (int, float)):
        if val < 0:
            return None
        if float(val) < 1:  # Excel fractional day
            total = int(round(float(val) * 24 * 60 * 60))  # seconds in the day
            if mmss_mode:
                h = total // 3600
                rem = total % 3600
                m = rem // 60
                # reinterpret h as minutes and m as seconds
                return h * 60 + m
            else:
                return float(total)
        return float(val)   # assume seconds

    # Datetime-like --------------------------------------------
    for attr in ("hour", "minute"):
        if hasattr(val, attr):
            h = int(getattr(val, "hour", 0))
            m = int(getattr(val, "minute", 0))
            s = int(getattr(val, "second", 0))
            if mmss_mode:
                # reinterpret h as minutes and m as seconds
                return h * 60 + m
            else:
                return h * 3600 + m * 60 + s

    return None


def _unwrap_rollover(seconds: pd.Series, threshold: float = 45.0, period: float = 3600.0) -> pd.Series:
    out = seconds.copy().astype(float)
    offset = 0.0
    prev = None
    for idx, v in out.items():
        if pd.isna(v):
            continue
        if prev is not None and (v + offset) < (prev - threshold):
            offset += period
        out.at[idx] = v + offset
        prev = out.at[idx]
    return out


def convert_time_to_seconds(series: pd.Series, anchor: str = "min", mmss_mode: bool = True) -> pd.Series:
    raw = series.apply(lambda x: _parse_time_to_seconds_raw(x, mmss_mode=mmss_mode))

    if raw.dropna().empty:
        return pd.Series([0]*len(series), index=series.index, dtype=float)

    raw_unwrapped = _unwrap_rollover(raw)

    start = raw_unwrapped.dropna().min() if anchor == "min" else raw_unwrapped.dropna().iloc[0]

    return raw_unwrapped.fillna(method="ffill").fillna(start).astype(float) - float(start)


def load_excel_data(file, anchor: str = "min", mmss_mode: bool = True) -> list[pd.DataFrame]:
    df = pd.read_excel(file, header=None)
    datasets = []
    for i in range(0, df.shape[1], 3):
        block = df.iloc[:, i:i+3]
        if block.shape[1] < 2:
            break
        block = block.rename(columns={block.columns[0]: "time",
                                      block.columns[1]: "pressure_1",
                                      block.columns[2] if block.shape[1] > 2 else block.columns[1]: "pressure_2"})
        block = block.dropna(how="all")
        if block.empty:
            continue
        block["time"] = convert_time_to_seconds(block["time"], anchor=anchor, mmss_mode=mmss_mode)
        block["pressure_1"] = pd.to_numeric(block["pressure_1"], errors="coerce")
        block["pressure_2"] = pd.to_numeric(block["pressure_2"], errors="coerce")
        block = block.dropna(subset=["pressure_1", "pressure_2"], how="all")
        datasets.append(block.reset_index(drop=True))
    return datasets


def calculate_axis_limits(datasets: list[pd.DataFrame], initial_pressure: float, auto: bool, manual: dict | None):
    if auto:
        xmin = float(min(d["time"].min() for d in datasets))
        xmax = float(max(d["time"].max() for d in datasets))
        allp = np.concatenate([pd.concat([d["pressure_1"].dropna(), d["pressure_2"].dropna()]).values for d in datasets])
        pmin = float(np.nanmin(allp))
        pmax = float(np.nanmax(allp))
        ypad = max(0.5, 0.05 * max(1.0, pmax - pmin))
        return {"x_min": xmin, "x_max": xmax, "y_min": pmin - ypad, "y_max": pmax + ypad}
    else:
        return {
            "x_min": float(manual.get("x_min", 0.0)),
            "x_max": float(manual.get("x_max", 300.0)),
            "y_min": float(manual.get("y_min", initial_pressure - 5)),
            "y_max": float(manual.get("y_max", initial_pressure + 5)),
        }


def create_plotly_figure(datasets: list[pd.DataFrame], limits: dict, names: list[str]):
    fig = go.Figure()
    palette = ["#e74c3c", "#3498db", "#e67e22", "#2ecc71", "#9b59b6", "#1abc9c"]
    for i, (data, name) in enumerate(zip(datasets, names)):
        color = palette[i % len(palette)]
        fig.add_trace(go.Scatter(x=data["time"], y=data["pressure_1"], mode="lines+markers",
                                 name=f"{name} - P1", marker={"size": 4}, line={"width": 2, "color": color}))
        fig.add_trace(go.Scatter(x=data["time"], y=data["pressure_2"], mode="lines+markers",
                                 name=f"{name} - P2", marker={"size": 4},
                                 line={"width": 2, "dash": "dot", "color": color}))
    fig.update_xaxes(title="Tempo (s)", range=[limits["x_min"], limits["x_max"]])
    fig.update_yaxes(title="Pressão (kPa)", range=[limits["y_min"], limits["y_max"]])
    fig.update_layout(height=560, legend={"orientation": "h"})
    return fig


def export_processed_data_excel(datasets, limits, initial_pressure, names):
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center = Alignment(horizontal="center", vertical="center")

    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Parâmetro", "Valor"])
    ws.append(["Pressão inicial (kPa)", initial_pressure])
    ws.append(["X min (s)", limits["x_min"]])
    ws.append(["X max (s)", limits["x_max"]])
    ws.append(["Y min (kPa)", limits["y_min"]])
    ws.append(["Y max (kPa)", limits["y_max"]])
    for c in ws[1]:
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    for name, data in zip(names, datasets):
        s = wb.create_sheet(name)
        cols = ["Tempo_s", "Pressao_1_kPa", "Pressao_2_kPa"]
        s.append(cols)
        for c in s[1]:
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
        for _, row in data.iterrows():
            s.append([float(row["time"]),
                      float(row["pressure_1"]) if not pd.isna(row["pressure_1"]) else None,
                      float(row["pressure_2"]) if not pd.isna(row["pressure_2"]) else None])

    comb = wb.create_sheet("Dados_Combinados")
    header = []
    for n in names:
        header += [f"Tempo_{n}_s", f"P1_{n}_kPa", f"P2_{n}_kPa"]
    comb.append(header)
    max_len = max(len(d) for d in datasets)
    for i in range(max_len):
        row = []
        for d in datasets:
            if i < len(d):
                row += [
                    float(d.loc[i, "time"]),
                    (None if pd.isna(d.loc[i, "pressure_1"]) else float(d.loc[i, "pressure_1"])),
                    (None if pd.isna(d.loc[i, "pressure_2"]) else float(d.loc[i, "pressure_2"])),
                ]
            else:
                row += [None, None, None]
        comb.append(row)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


def export_processed_data_csv(datasets, limits, initial_pressure, names):
    s = io.StringIO()
    s.write(f"# Dados Ajustados - Pressão Inicial: {initial_pressure} kPa\n")
    s.write(f"# X: {limits['x_min']} .. {limits['x_max']} s | Y: {limits['y_min']} .. {limits['y_max']} kPa\n")
    s.write("#\n")
    header = []
    for n in names:
        header += [f"Tempo_{n}_s", f"P1_{n}_kPa", f"P2_{n}_kPa"]
    s.write(",".join(header) + "\n")
    max_len = max(len(d) for d in datasets)
    for i in range(max_len):
        row = []
        for d in datasets:
            if i < len(d):
                row += [
                    f"{float(d.loc[i, 'time']):.6f}",
                    "" if pd.isna(d.loc[i, 'pressure_1']) else f"{float(d.loc[i, 'pressure_1']):.6f}",
                    "" if pd.isna(d.loc[i, 'pressure_2']) else f"{float(d.loc[i, 'pressure_2']):.6f}",
                ]
            else:
                row += ["", "", ""]
        s.write(",".join(row) + "\n")
    return s.getvalue().encode("utf-8")


def export_alignment_parameters(datasets, limits, initial_pressure, names):
    stats = []
    for n, d in zip(names, datasets):
        stats.append({
            "name": n,
            "points": int(len(d)),
            "time_range": {"min": float(d["time"].min()), "max": float(d["time"].max())},
            "pressure_1": {
                "min": float(np.nanmin(d["pressure_1"])), "max": float(np.nanmax(d["pressure_1"])),
                "mean": float(np.nanmean(d["pressure_1"])), "std": float(np.nanstd(d["pressure_1"]))},
            "pressure_2": {
                "min": float(np.nanmin(d["pressure_2"])), "max": float(np.nanmax(d["pressure_2"])),
                "mean": float(np.nanmean(d["pressure_2"])), "std": float(np.nanstd(d["pressure_2"]))},
        })
    payload = {
        "exported_at": datetime.now().isoformat(timespec="seconds"),
        "initial_pressure_kPa": float(initial_pressure),
        "limits": {
            "x_min_s": float(limits["x_min"]), "x_max_s": float(limits["x_max"]),
            "y_min_kPa": float(limits["y_min"]), "y_max_kPa": float(limits["y_max"]),
        },
        "datasets": stats,
    }
    return json.dumps(payload, indent=2, ensure_ascii=False).encode("utf-8")


# -----------------------------
# UI
# -----------------------------

def main():
    st.set_page_config(page_title="Função Degrau - Pressão", layout="wide")

    st.title("📈 Função Degrau — Gráfico de Pressão")
    st.write("Carregue um Excel com colunas em trincas: tempo, pressão 1, pressão 2 (para cada dataset).")

    with st.sidebar:
        uploaded = st.file_uploader("Arquivo Excel (.xlsx)", type=["xlsx"])
        fmt = st.radio("Formato do tempo no Excel", ["mm:ss (mesmo se aparecer HH:MM:SS)", "hh:mm:ss real"], index=0)
        mmss_mode = (fmt.startswith("mm:ss"))
        anchor = st.radio("Ancoragem do tempo", ["menor valor (recomendado)", "primeiro valor"], index=0)
        initial_pressure = st.number_input("Pressão inicial (kPa)", value=50.0, step=0.5)
        auto = st.checkbox("Ajustar limites automaticamente", value=True)
        manual = {}
        if not auto:
            manual["x_min"] = st.number_input("X mín (s)", value=0.0)
            manual["x_max"] = st.number_input("X máx (s)", value=300.0)
            manual["y_min"] = st.number_input("Y mín (kPa)", value=initial_pressure - 5.0)
            manual["y_max"] = st.number_input("Y máx (kPa)", value=initial_pressure + 5.0)
        debug = st.checkbox("Modo debug (mostrar infos do tempo)", value=False)

    if uploaded is None:
        st.info("Faça o upload do arquivo para visualizar o gráfico.")
        return

    try:
        datasets = load_excel_data(uploaded, anchor=("min" if anchor.startswith("menor") else "first"), mmss_mode=mmss_mode)
        if not datasets:
            st.warning("Não encontrei colunas suficientes (grupos de 3). Verifique o arquivo.")
            return
    except Exception as e:
        st.error(f"Erro ao ler o Excel: {e}")
        return

    names = [f"Dataset {i+1}" for i in range(len(datasets))]
    limits = calculate_axis_limits(datasets, initial_pressure, auto, manual if not auto else None)
    fig = create_plotly_figure(datasets, limits, names)
    st.plotly_chart(fig, use_container_width=True)

    # Exports
    col1, col2, col3 = st.columns(3)
    with col1:
        xlsx_bytes = export_processed_data_excel(datasets, limits, initial_pressure, names)
        st.download_button("⬇️ Baixar Excel", data=xlsx_bytes, file_name="dados_ajustados.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    with col2:
        csv_bytes = export_processed_data_csv(datasets, limits, initial_pressure, names)
        st.download_button("⬇️ Baixar CSV", data=csv_bytes, file_name="dados_ajustados.csv", mime="text/csv")
    with col3:
        json_bytes = export_alignment_parameters(datasets, limits, initial_pressure, names)
        st.download_button("⬇️ Baixar parâmetros (JSON)", data=json_bytes, file_name="parametros.json",
                           mime="application/json")

    if debug:
        st.subheader("Debug do Tempo (primeiras linhas de cada dataset)")
        for i, d in enumerate(datasets, start=1):
            st.write(f"**Dataset {i}** — {len(d)} pontos")
            st.write(d.head(10))

if __name__ == "__main__":
    main()
